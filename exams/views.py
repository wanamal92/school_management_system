# exams/views.py
import matplotlib.pyplot as plt
from django.shortcuts import render, redirect, get_object_or_404
from .models import ExamSession, Exam, ExamAttendee
from .forms import ExamSessionForm, ExamForm, ExamAttendeeForm, ExcelUploadForm
import pandas as pd
from clases.models import Class
from students.models import Student
from django.db import IntegrityError
from django.core.exceptions import ValidationError
from zipfile import BadZipFile
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from django.db import IntegrityError
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.conf import settings
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
import os
from django.contrib import messages
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from django.http import HttpResponse
from .models import ExamSession, Exam, ExamAttendee, Student
from reportlab.lib.utils import ImageReader
import matplotlib
matplotlib.use('Agg')

def is_admin(user):
    return user.is_authenticated and user.role == 'admin'

def is_admin_or_staff(user):
    return user.is_authenticated and user.role in ['admin', 'staff']


# Exam sessions
@user_passes_test(is_admin_or_staff)
@login_required
def list_exam_session(request):
    exam_sessions = ExamSession.objects.all().order_by('exam_session_name')

    # Pagination setup
    paginator = Paginator(exam_sessions, 100)  # Show 10 students per page
    page_number = request.GET.get('page')
    exam_sessions_page = paginator.get_page(page_number)

    return render(request, 'exams/list_exam_session.html', {'exam_sessions': exam_sessions_page})

@user_passes_test(is_admin_or_staff)
@login_required
def create_exam_session(request):
    if request.method == 'POST':
        form = ExamSessionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam Session created successfully!")
            return redirect('list_exam_session')
        else:
            messages.error(request, "There was an error in creating the exam session.")        
    else:
        form = ExamSessionForm()
    return render(request, 'exams/create_exam_session.html', {'form': form})

@user_passes_test(is_admin_or_staff)
@login_required
def edit_exam_session(request, pk):
    exam_session = get_object_or_404(ExamSession, pk=pk)
    if request.method == 'POST':
        form = ExamSessionForm(request.POST, instance=exam_session)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam Session updated successfully!")            
            return redirect('list_exam_session')
        else:
            messages.error(request, "There was an error updating the exam session.")        
    else:
        form = ExamSessionForm(instance=exam_session)
    return render(request, 'exams/create_exam_session.html', {'form': form})

@user_passes_test(is_admin)
@login_required
def delete_exam_session(request, pk):
    exam_session = get_object_or_404(ExamSession, pk=pk)
    if request.method == 'POST':
        exam_session.delete()
        messages.success(request, "Exam Session deleted successfully!")
        return redirect('list_exam_session')
    return redirect('list_exam_session')

@user_passes_test(is_admin_or_staff)
@login_required
def detail_exam_session(request, pk):
    exam_session = get_object_or_404(ExamSession, pk=pk)
    return render(request, 'exams/detail_exam_session.html', {'exam_session': exam_session})


# Exams
@user_passes_test(is_admin_or_staff)
@login_required
def list_exam(request):
    exams = Exam.objects.all().order_by('exam_name')

    # Pagination setup
    paginator = Paginator(exams, 100)  # Show 10 students per page
    page_number = request.GET.get('page')
    exams_page = paginator.get_page(page_number)

    return render(request, 'exams/list_exam.html', {'exams': exams_page})

@user_passes_test(is_admin_or_staff)
@login_required
def create_exam(request):
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam created successfully!")
            return redirect('list_exam')
        else:
            messages.error(request, "There was an error in creating the exam.")
        
    else:
        form = ExamForm()
    return render(request, 'exams/create_exam.html', {'form': form})

@user_passes_test(is_admin_or_staff)
@login_required
def edit_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam updated successfully!")
            return redirect('list_exam')
        else:
            messages.error(request, "There was an error updating the exam.")
    else:
        form = ExamForm(instance=exam)
    return render(request, 'exams/create_exam.html', {'form': form})

@user_passes_test(is_admin)
@login_required
def delete_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        exam.delete()
        messages.success(request, "Exam deleted successfully!")
        return redirect('list_exam')
    return redirect('list_exam')

@user_passes_test(is_admin_or_staff)
@login_required
def detail_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    return render(request, 'exams/detail_exam.html', {'exam': exam})


# Exam Attendees
@user_passes_test(is_admin_or_staff)
@login_required
def list_exam_attendee(request):
    exam_attendees = ExamAttendee.objects.all().order_by('student__full_name')

    # Pagination setup
    paginator = Paginator(exam_attendees, 200)  # Show 10 students per page
    page_number = request.GET.get('page')
    exam_attendees_page = paginator.get_page(page_number)

    return render(request, 'exams/list_exam_attendee.html', {'exam_attendees': exam_attendees_page})

@user_passes_test(is_admin_or_staff)
@login_required
def create_exam_attendee(request):
    if request.method == 'POST':
        form = ExamAttendeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam Attendee created successfully!")
            return redirect('list_exam_attendee')
        else:
            messages.error(request, "There was an error in creating the exam attendee.")
    else:
        form = ExamAttendeeForm()
    return render(request, 'exams/create_exam_attendee.html', {'form': form})

@user_passes_test(is_admin_or_staff)
@login_required
def edit_exam_attendee(request, pk):
    exam_attendee = get_object_or_404(ExamAttendee, pk=pk)
    if request.method == 'POST':
        form = ExamAttendeeForm(request.POST, instance=exam_attendee)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam Attendee updated successfully!")
            return redirect('list_exam_attendee')
        else:
            messages.error(request, "There was an error updating the exam attendee.")
    else:
        form = ExamAttendeeForm(instance=exam_attendee)
    return render(request, 'exams/create_exam_attendee.html', {'form': form})

@user_passes_test(is_admin)
@login_required
def delete_exam_attendee(request, pk):
    exam_attendee = get_object_or_404(ExamAttendee, pk=pk)
    if request.method == 'POST':
        exam_attendee.delete()
        messages.success(request, "Exam Attendee deleted successfully!")
        return redirect('list_exam_attendee')
    return redirect('list_exam_attendee')

@user_passes_test(is_admin_or_staff)
@login_required
def detail_exam_attendee(request, pk):
    exam_attendee = get_object_or_404(ExamAttendee, pk=pk)

    return render(request, 'exams/detail_exam_attendee.html', {'exam_attendee': exam_attendee})

@user_passes_test(is_admin_or_staff)
@login_required
def upload_exam_session_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the uploaded Excel file
            excel_file = request.FILES['excel_file']

            # Check if the file is a valid Excel file (by trying to read it)
            try:
                # Try reading the Excel file (forcing pandas to use the right engine for .xlsx)
                # Specify engine for .xlsx files
                df = pd.read_excel(excel_file, engine='openpyxl')

            except BadZipFile:
                # Handle the error for invalid Excel files (not a valid ZIP/Excel file)
                return render(request, 'exams/upload_exam_session_excel.html', {
                    'form': form,
                    'error': 'Error: The uploaded file is not a valid Excel file. Please upload a valid Excel file.'
                })

            except ValueError as e:
                # If pandas cannot read the file, show a friendly error
                return render(request, 'exams/upload_exam_session_excel.html', {
                    'form': form,
                    'error': 'Error: Invalid file format. Please upload a valid Excel file.'
                })

            # Proceed with the rest of the file processing (create records, etc.)
            for index, row in df.iterrows():
                # Fetch the Class instance based on the class name (e.g., 'Grade 1')
                try:
                    class_assigned = Class.objects.get(
                        class_name=row['class_assigned'])
                except Class.DoesNotExist:
                    class_assigned = None  # Handle case where class doesn't exist

                # Check if exam_session_code already exists
                if ExamSession.objects.filter(exam_session_code=row['exam_session_code']).exists():
                    # Return error if the session code already exists
                    return render(request, 'exams/upload_exam_session_excel.html', {
                        'form': form,
                        'error': f"Error: Exam session code '{row['exam_session_code']}' already exists!"
                    })

                try:
                    # Create ExamSession if no error
                    ExamSession.objects.create(
                        exam_session_name=row['exam_session_name'],
                        exam_session_code=row['exam_session_code'],
                        start_date=row['start_date'],
                        end_date=row['end_date'],
                        exam_type=row['exam_type'],
                        academic_year=row['academic_year'],
                        class_assigned=class_assigned,
                    )
                except IntegrityError:
                    return render(request, 'exams/upload_exam_session_excel.html', {
                        'form': form,
                        'error': f"Error: Integrity error occurred for row {index + 1}. Please check your data."
                    })

            # Redirect after successful upload
            return redirect('list_exam_sessions')
    else:
        form = ExcelUploadForm()

    return render(request, 'exams/upload_exam_session_excel.html', {'form': form})


def get_grade(marks):
    if marks >= 75:
        return 'A'
    elif marks >= 65:
        return 'B'
    elif marks >= 50:
        return 'C'
    elif marks >= 35:
        return 'S'
    else:
        return 'W'

@user_passes_test(is_admin_or_staff)
@login_required
def export_exam_session_excel(request, session_id):
    session = ExamSession.objects.get(id=session_id)
    exams = Exam.objects.filter(
        exam_session=session).select_related('exam_name')
    students = Student.objects.filter(class_level=session.class_assigned)

    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Report"

    total_columns = 1 + exams.count() + 3  # Student + subjects + Total, Avg, Rank

    # Add logo
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')
    if os.path.exists(logo_path):
        logo_img = XLImage(logo_path)
        logo_img.width = 50
        logo_img.height = 50
        ws.add_image(logo_img, "A1")

    # Heading
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=total_columns)
    title_cell = ws.cell(
        row=1, column=1, value="Sri Gnanalankara Maha Pirivena - Peradeniya")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Session info
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=total_columns)
    session_title = f"Exam Session: {session.exam_session_name} ({session.academic_year}) "
    session_cell = ws.cell(row=2, column=1, value=session_title)
    session_cell.font = Font(size=12, bold=True)
    session_cell.alignment = Alignment(horizontal='center')

    # Table headers
    headers = ['Student Name'] + [exam.exam_name.subject_name for exam in exams] + \
        ['Total Marks', 'Average', 'Rank']
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C0C0C0",
                                end_color="C0C0C0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Max marks row
    max_marks_row = ['Max Marks'] + ['100'] * \
        exams.count() + [100*exams.count(), '', '']
    ws.append(max_marks_row)
    for col_num, val in enumerate(max_marks_row, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="E0F7FA",
                                end_color="E0F7FA", fill_type="solid")

    # Student rows
    data_rows = []
    for student in students:
        row = [student.full_name]
        total = 0
        count = 0
        for exam in exams:
            attendee = ExamAttendee.objects.filter(
                student=student, exam=exam).first()
            if attendee and attendee.status == 'present':
                mark = attendee.exam_marks
                grade = get_grade(mark)
                row.append(f"{mark} ({grade})")
                total += mark
                count += 1
            elif attendee:
                row.append("Absent")
            else:
                row.append("N/A")
        avg = total / count if count else 0
        row.append(total)
        row.append(round(avg, 2))
        row.append(0)  # Rank placeholder
        data_rows.append(row)

    data_rows.sort(key=lambda x: x[-3], reverse=True)
    for idx, row in enumerate(data_rows):
        row[-1] = idx + 1
        ws.append(row)

    # Adjust column widths
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    # Conditional formatting
    red_fill = PatternFill(start_color="FFCDD2",
                           end_color="FFCDD2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4",
                              end_color="FFF9C4", fill_type="solid")
    red_style = DifferentialStyle(fill=red_fill)
    yellow_style = DifferentialStyle(fill=yellow_fill)
    start_row = 5
    end_row = 4 + len(data_rows)
    for col in range(2, 2 + len(exams)):
        col_letter = get_column_letter(col)
        red_rule = Rule(type="containsText",
                        operator="containsText", text="(W)", dxf=red_style)
        yellow_rule = Rule(
            type="containsText", operator="containsText", text="(S)", dxf=yellow_style)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}", red_rule)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}", yellow_rule)

    # Chart
    chart = BarChart()
    chart.title = "Total Marks by Student"
    chart.y_axis.title = "Total Marks"
    chart.x_axis.title = "Students"
    chart.style = 12
    chart.height = 10
    chart.width = 20
    data_ref = Reference(ws, min_col=len(
        headers)-2, max_col=len(headers)-2, min_row=5, max_row=4+len(data_rows))
    cats_ref = Reference(ws, min_col=1, max_col=1,
                         min_row=5, max_row=4+len(data_rows))
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    ws.add_chart(chart, f"A{len(data_rows) + 7}")

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{session.exam_session_name}_report.xlsx"'
    return response

@user_passes_test(is_admin_or_staff)
@login_required
def export_exam_session_pdf(request, session_id):
    session = ExamSession.objects.get(id=session_id)
    exams = Exam.objects.filter(
        exam_session=session).select_related('exam_name')
    students = Student.objects.filter(class_level=session.class_assigned)

    # Headers: subject names + total, average, rank
    headers = ['Student Name'] + \
        [exam.exam_name.subject_name for exam in exams] + \
        ['Total', 'Average', 'Rank']
    data_rows = []

    for student in students:
        row = [student.full_name]
        total = 0
        count = 0
        for exam in exams:
            attendee = ExamAttendee.objects.filter(
                student=student, exam=exam).first()
            if attendee and attendee.status == 'present':
                row.append(attendee.exam_marks)
                total += attendee.exam_marks
                count += 1
            elif attendee:
                row.append("Absent")
            else:
                row.append("N/A")
        avg = total / count if count else 0
        row.append(total)
        row.append(round(avg, 2))
        row.append(0)  # Temporary rank
        data_rows.append(row)

    # Sort and rank
    data_rows.sort(key=lambda x: x[-3], reverse=True)
    for i, row in enumerate(data_rows):
        row[-1] = i + 1

    table_data = [headers] + data_rows

    # Generate PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{session.exam_session_name}_report.pdf"'
    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, height - 40,
                        f"Exam Report - {session.exam_session_name} ({session.academic_year})")

    # Draw Table
    table = Table(table_data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(style)

    # Position the table
    table_width, table_height = table.wrap(0, 0)
    x = (width - table_width) / 2
    y = height - 100 - table_height
    table.wrapOn(c, width, height)
    table.drawOn(c, x, y)

    # ===== CHART GENERATION =====
    student_names = [row[0] for row in data_rows]
    total_marks = [row[-3] for row in data_rows]

    plt.figure(figsize=(10, 4))
    bars = plt.bar(student_names, total_marks, color='skyblue')
    plt.xlabel("Students")
    plt.ylabel("Total Marks")
    plt.title("Total Marks by Student")
    plt.xticks(rotation=45, ha='right')

    # Add value labels on top of bars
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 1,
                 f'{height}', ha='center', va='bottom', fontsize=8)

    # Save chart to in-memory image
    img_buffer = BytesIO()
    plt.tight_layout()
    plt.savefig(img_buffer, format='PNG')
    plt.close()
    img_buffer.seek(0)
    chart_img = ImageReader(img_buffer)

    # Draw chart image below the table
    chart_width = 500
    chart_height = 250
    c.drawImage(chart_img, x=50, y=50, width=chart_width, height=chart_height)

    # Finalize PDF
    c.showPage()
    c.save()
    return response


@login_required
def schedule_exam_attendees(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    exam_session = exam.exam_session
    assigned_class = exam_session.class_assigned
    students = Student.objects.filter(class_level=assigned_class)

    created_count = 0
    skipped_count = 0

    for student in students:
        try:
            obj, created = ExamAttendee.objects.get_or_create(
                student=student,
                exam=exam,
                defaults={'exam_marks': 0, 'status': 'present'}
            )
            if created:
                created_count += 1
            else:
                skipped_count += 1
        except IntegrityError:
            skipped_count += 1  # just in case race condition or violation

    if created_count == len(students):
        messages.success(
            request, f"Successfully scheduled all {created_count} students for the exam.")
    elif created_count > 0:
        messages.warning(
            request, f"Scheduled {created_count} students. {skipped_count} students were already scheduled.")
    else:
        messages.info(
            request, "All students were already scheduled for this exam.")

    return redirect('detail_exam', pk=exam.id)
