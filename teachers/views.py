# teachers/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import Teacher
from leaves.models import LeaveRequest
from .forms import TeacherForm
from django.contrib.auth.decorators import login_required, user_passes_test
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
import os
from django.db import models
from leaves.models import LeaveType, LeaveRequest, LeaveAllocation
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.contrib import messages
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
from django.http import HttpResponse
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib import rcParams
rcParams.update({'figure.autolayout': True})


def is_admin(user):
    return user.is_authenticated and user.role == 'admin'

# List all teachers


@user_passes_test(is_admin)
@login_required
def list_teachers(request):
    # Get the view mode from the URL parameter (default is 'list')
    view_mode = request.GET.get('view', 'list')

    # Fetch all the teachers (you can modify this query based on your needs)
    teachers = Teacher.objects.all().order_by('full_name')

    # Pagination setup
    paginator = Paginator(teachers, 10)  # Show 10 students per page
    page_number = request.GET.get('page')
    teachers_page = paginator.get_page(page_number)

    # Pass the teachers and view_mode to the template
    return render(request, 'teachers/list_teachers.html', {
        'teachers': teachers_page,
        'view_mode': view_mode,
    })

# Add a new teacher


@user_passes_test(is_admin)
@login_required
def add_teacher(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('list_teachers')
    else:
        form = TeacherForm()
    return render(request, 'teachers/add_teacher.html', {'form': form})

# Edit an existing teacher


@user_passes_test(is_admin)
@login_required
def edit_teacher(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            return redirect('list_teachers')
    else:
        form = TeacherForm(instance=teacher)
    return render(request, 'teachers/edit_teacher.html', {'form': form})

# Delete a teacher


@user_passes_test(is_admin)
@login_required
def delete_teacher(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    try:
        if request.method == 'POST':
            teacher.delete()
            return redirect('list_teachers')
    except IntegrityError as e:
        # Handling IntegrityError (ForeignKey constraint violation)
        if 'foreign key' in str(e).lower():
            messages.error(
                request, "Cannot delete this user because they are referenced in other records (e.g., as class in charge).")
        else:
            messages.error(request, f"A database error occurred: {str(e)}")
        return redirect('list_teachers')

    return redirect('list_teachers')


@user_passes_test(is_admin)
@login_required
def detail_teacher(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    return render(request, 'teachers/detail_teacher.html', {'teacher': teacher})


@login_required
def profile_teacher(request):
    # Fetch the teacher's profile related to the logged-in user
    try:
        teacher_profile = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        teacher_profile = None  # or handle if the profile does not exist

    return render(request, 'teachers/profile_teacher.html', {'teacher_profile': teacher_profile})


@login_required
def export_teacher_leave_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    leave_types = LeaveType.objects.all()
    # Name + (Used + Remaining) per leave type
    total_columns = 1 + leave_types.count() * 2

    # === Logo ===
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')
    if os.path.exists(logo_path):
        logo_img = XLImage(logo_path)
        logo_img.width = 50
        logo_img.height = 50
        ws.add_image(logo_img, "A1")

    # === Main Heading ===
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=total_columns)
    main_title_cell = ws.cell(
        row=1, column=1, value="Sri Gnanalankara Maha Pirivena - Peradeniya")
    main_title_cell.font = Font(size=16, bold=True)
    main_title_cell.alignment = Alignment(horizontal='center')

    # === Sub Heading ===
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=total_columns)
    sub_title_cell = ws.cell(row=2, column=1, value="Teachers' Leaves Records")
    sub_title_cell.font = Font(size=14, bold=True)
    sub_title_cell.alignment = Alignment(horizontal='center')

    # === Table Headers ===
    headers = ['Teacher Name']
    for lt in leave_types:
        headers.append(f"{lt.name} - Used")
        headers.append(f"{lt.name} - Remaining")
    ws.append(headers)

    # Style headers
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C0C0C0",
                                end_color="C0C0C0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # === Fill Data and track total used leaves per teacher ===
    teacher_rows = []  # Will store (row_number, total_used) for charting
    for idx, teacher in enumerate(Teacher.objects.all(), start=4):
        row = [teacher.full_name]
        total_used = 0
        allocation = LeaveAllocation.objects.filter(teacher=teacher).first()

        for lt in leave_types:
            used = LeaveRequest.objects.filter(
                teacher=teacher, leave_type=lt, status='Approved'
            ).aggregate(models.Sum('duration'))['duration__sum'] or 0

            if allocation:
                field_name = lt.name.lower().replace(' ', '_')
                remaining = getattr(allocation, field_name, 0)
            else:
                remaining = 0

            row.extend([used, remaining])
            total_used += used

        ws.append(row)
        teacher_rows.append((idx, total_used))

    # === Adjust column widths ===
    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(
            col[0].column)].width = max_length + 2

    # === Add Bar Chart for Total Used Leaves per Teacher ===
    chart = BarChart()
    chart.type = "bar"  # horizontal bar chart
    chart.style = 13    # nicer style
    chart.title = "Total Leaves Used by Teacher"
    chart.y_axis.title = "Teachers"
    chart.x_axis.title = "Leaves Used"
    chart.height = 8
    chart.width = 14

    start_data_row = 4  # Data starts from row 4 (after header)
    num_teachers = len(teacher_rows)

    # Add 'Total Used' column after existing columns for chart source
    total_used_col = total_columns + 1
    ws.cell(row=3, column=total_used_col,
            value='Total Used').font = Font(bold=True)
    ws.column_dimensions[get_column_letter(total_used_col)].width = 14

    for row_num, total_used in teacher_rows:
        ws.cell(row=row_num, column=total_used_col, value=total_used)

    data_ref = Reference(ws, min_col=total_used_col, max_col=total_used_col,
                         min_row=start_data_row, max_row=start_data_row + num_teachers - 1)
    cats_ref = Reference(ws, min_col=1, max_col=1,
                         min_row=start_data_row, max_row=start_data_row + num_teachers - 1)

    # === Save and Return Excel Response ===
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="teachers_leave_records.xlsx"'}
    )


@login_required
def export_teacher_leave_pdf(request):
    # Gather teacher leave data
    allocations = LeaveAllocation.objects.select_related('teacher')
    leave_requests = LeaveRequest.objects.filter(status='Approved')

    data_rows = []
    for allocation in allocations:
        teacher = allocation.teacher
        casual_taken = leave_requests.filter(teacher=teacher, leave_type__name='Casual Leave').aggregate(
            total=models.Sum('duration'))['total'] or 0
        sick_taken = leave_requests.filter(teacher=teacher, leave_type__name='Sick Leave').aggregate(
            total=models.Sum('duration'))['total'] or 0

        data_rows.append([
            teacher.full_name,
            casual_taken,
            allocation.casual_leave,
            sick_taken,
            allocation.sick_leave,
            casual_taken + sick_taken,
        ])

    # Sort by total leave taken
    data_rows.sort(key=lambda x: x[-1], reverse=True)

    # PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="teacher_leave_report.pdf"'
    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, height - 40,
                        "Sri Gnanalankara Maha Pirivena - Peradeniya")
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, height - 65, "Teachers' Leaves Records")

    # Table headers
    headers = ['Teacher Name', 'Casual Taken', 'Casual Left',
               'Sick Taken', 'Sick Left', 'Total Taken']
    table_data = [headers] + data_rows

    # Table
    table = Table(table_data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(style)

    table_width, table_height = table.wrap(0, 0)
    x = (width - table_width) / 2
    y = height - 120 - table_height
    table.wrapOn(c, width, height)
    table.drawOn(c, x, y)

    # Chart: Bar chart of total leaves taken
    teacher_names = [row[0] for row in data_rows]
    total_taken = [row[-1] for row in data_rows]

    plt.figure(figsize=(10, 4))
    bars = plt.bar(teacher_names, total_taken, color='salmon')
    plt.title('Total Leave Taken by Teachers')
    plt.ylabel('Days')
    plt.xticks(rotation=45, ha='right', fontsize=8)

    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                 f'{height}', ha='center', va='bottom', fontsize=7)

    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='PNG')
    plt.close()
    img_buffer.seek(0)
    chart_img = ImageReader(img_buffer)

    c.drawImage(chart_img, x=50, y=50, width=500, height=220)

    # Finish
    c.showPage()
    c.save()
    return response
